from openpyxl import load_workbook

wb2 = load_workbook('CM2_1.xlsx')

print (wb2.sheetnames)

ws = wb2.active
print("Customers")
for row in ws.values:
    for value in row:
        print(value)

print("Invoices")
ws = wb2["Invoices"]
for row in ws.values:
    for value in row:
        print(value)

print("Line Items")
ws = wb2["Line Items"]
for row in ws.values:
    for value in row:
        print(value)

print("Products")
ws = wb2["Products"]
for row in ws.values:
    for value in row:
        print(value)