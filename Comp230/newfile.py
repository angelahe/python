from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet("Mysheet")

ws.title = "Contact"

print(wb.sheetnames)

ws['A4'] = 4

d =ws.cell(row=4, column=2, value=10)

for row in ws.values:
    for value in row:
        print(value)

wb.save('testing.xlsx')

