from openpyxl import Workbook

def make_invoice_header(ws):
    ws['A1'] = "BILL'S CONSULTING CORP"
    ws['A2'] = "123 Acorn St NW"
    ws['A3'] = "Canoe, AB"
    ws['A4'] = "Canada T0M 0R0"
    ws['A5'] = "555-555-1234"

def make_invoice_billing_section(ws):
    ws['A8'] = "BILL TO"
    ws['A9'] = "Glenna Mann"
    ws['A10'] = "122 Abby St SW Canmore AB"
    ws['A11'] = "111-111-1234"
    ws['A12'] = "glenna@mannservices.com"

def make_invoice_id_section(ws, invoice_id, date):
    ws['F1'] = "INVOICE"
    ws['F3'] = "DATE"
    ws['H3'] = "INVOICE #"
    ws['F4'] = date
    ws['H4'] = invoice_id

def make_invoice_items_section(ws):
    ws['A17'] = "DESCRIPTION"
    ws['F17'] = "QTY"
    ws['G17'] = "COST"
    ws['H17'] = "AMOUNT"

#create lineitem in spreadsheet and return lineitem total
def make_invoice_line_item(ws, item, qty, price, line_number):
    linenum = str(line_number)
    ws['A'+linenum] = item
    ws['F'+linenum] = qty
    ws['G'+linenum] = price
    ws['H'+linenum] = qty*price
    return(qty*price)

def make_invoice_total_section(ws, total, line_number):
    linenum = str(line_number)
    ws['F'+linenum] = "TOTAL"
    ws['H'+linenum] = total

def make_invoice(invoice_id):

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    make_invoice_header(ws)
    make_invoice_billing_section(ws)
    make_invoice_id_section(ws, invoice_id, "10/26/2019")
    make_invoice_items_section(ws)
    
    next_line_number = 18
    total = 0
    total+=make_invoice_line_item(ws, "Apples", 10, 1.25, next_line_number)
    next_line_number+=1
    total+=make_invoice_line_item(ws, "Bananas", 1, 2.00, next_line_number)
    next_line_number+=2

    make_invoice_total_section(ws, total, next_line_number)

    wb.save(invoice_id+'.xlsx')


make_invoice('CDN10107')
