from openpyxl import load_workbook

class GetInvoiceInfo():
    
    def __init__(self, invoiceID, fileName):
        self.invoiceID = invoiceID
        self.fileName = fileName
        self.customerID = 0
        self.invoiceDate = ''
        self.customerName = ''
        self.customerAddress = ''
        self.customerPhone = ''
        self.customerContact = ''
        self.invoiceItems = []
        self.products = []

    def check_file_exists(self):
        '''
            check the file can open successfully
        '''
        try:
            with open(self.fileName, 'r') as fh:
                return True
        except FileNotFoundError:
            print('file not found')
            return False

    def check_for_sheets(self):
        '''
            check that Customers, Invoices, Line Items, and Products sheets 
            are in the file
        '''
        wb = load_workbook(self.fileName)
        sheetnames = wb.sheetnames
        neededSheets = ['Customers', 'Invoices', 'Line Items', 'Products']
        for sheet in neededSheets:
            if sheet not in sheetnames :
                print (f'Missing required {sheet} sheet in workbook')
                return False
        return True

    def get_invoice_info(self):
        '''
            lookup the provided invoiceID, error if not found
            assume columns in sheet are InvoiceID, CustomerID, Date
            load the customerID and invoiceDate into class attributes
            return true if found, else false
        '''
        wb = load_workbook(self.fileName)
        ws = wb['Invoices']
        for row in range(2, ws.max_row+1):
            if (ws['A'+str(row)].value == self.invoiceID):
                self.customerID = ws['B'+str(row)].value
                self.invoiceDate = ws['C'+str(row)].value
                return True
        print(f'invoice {self.invoiceID} was not found')
        return False

    def get_customer_info(self):
        '''
            lookup the customer info using the customerID from the Invoice sheet
            assume columns are CusomerID, Customer, Address, Phone, Contact
            return true if found, else false
        '''
        wb = load_workbook(self.fileName)
        if self.customerID == 0 :
            if self.get_invoice_info() == False :
                print('No customer for this invoice')
                return False

        ws = wb['Customers']
        for row in range(2, ws.max_row+1):
            if (ws['A'+str(row)].value == self.customerID):
                self.customerName = ws['B'+str(row)].value
                self.customerAddress = ws['C'+str(row)].value
                self.customerPhone = ws['D'+str(row)].value
                self.customerContact = ws['E'+str(row)].value
                return True
        print(f'customer {self.customerID} was not found')
        return False

    def get_line_items(self):
        '''
            get all the line items for the provided invoiceID
            return false if no items found
        '''
        wb = load_workbook(self.fileName)
        if self.products == []:
            if self.load_product_info() == False:
                print('Could not find products')
                return False

        ws = wb['Line Items']
        count = 0
        #find the line items with the defined invoice#
        for row in range(2, ws.max_row + 1):
            if ws['A'+str(row)].value == self.invoiceID :
                print('found the invoice')
                count+=1
                productNum = ws['B'+str(row)].value
                productInfo = self.lookup_product_info(productNum)
                units = ws['C'+str(row)].value
                lineitem = { 
                    'productNum': productNum,
                    'description': productInfo['description'],
                    'unitCost': productInfo['unitCost'],
                    'units': units,
                    'totalCost': productInfo['unitCost'] * units
                }
                self.invoiceItems.append(lineitem)
        #return false if no line items are found for the invoice
        if count == 0:
            print(f'No item found for invoice {self.invoiceID}')
            return False
        else:
            #print('these are the items')
            #print(self.invoiceItems)
            return True

    def load_product_info(self):
        '''
        load all the product info for lookup
        assume columns in order are ProductID, Description, Unit Cost
        returns a dictionary of all product info
        '''
        wb = load_workbook(self.fileName)
        ws = wb['Products']
        count = 0
        for row in range(2, ws.max_row+1):
            count+=1
            productNum = ws['A'+str(row)].value
            description = ws['B'+str(row)].value
            unitCost = ws['C'+str(row)].value
            productInfo = {
                'productNum': productNum, 
                'description': description,
                'unitCost': unitCost
            }
            self.products.append(productInfo)
        if count == 0:
            print('No products found')
            return False
        else:
            #print('the products are')
            #print(self.products)
            return True

    def lookup_product_info(self, productID):
        '''
        given a productID look up the info for the product
        return False if no product was found for the productID
        '''
        if self.products == [] :
            print('No products found')
            return False
        
        result = next((item for item in self.products if item["productNum"] == productID), False)
        
        print(f"result is {result}")
        return result

