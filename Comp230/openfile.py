from openpyxl import load_workbook

wb2 = load_workbook('CM.xlsx')

print (wb2.sheetnames)

ws = wb2.active
print("Sheet 1")
for row in ws.values:
    for value in row:
        print(value)

print("Sheet 2")
ws = wb2["Sheet 2"]
for row in ws.values:
    for value in row:
        print(value)

print("Sheet 3")
ws = wb2["Sheet 3"]
for row in ws.values:
    for value in row:
        print(value)